from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException,  TimeoutException

import re
from bs4 import BeautifulSoup
import time
import requests
import random
from random import randint
from time import sleep
from openpyxl import Workbook, load_workbook
import os
import logging


def create_workbook():
    # This function creates an Excel workbook and check the condition whether the file is present or not,
    # if file is not present it will create it or else it will append it.

    excel_file_name = "Reddit Data.xlsx"
    file_path = os.getcwd() + '\\' + excel_file_name
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active

        # Below three lines are written due to in some circumstances if the user deletes the headers in the file and then runs the code, This logic will place the headers again.

        if sheet['A1'].value is None:
            sheet.append(['Title', 'Author', 'UpVotes', 'Number of Comments', 'Post URl', 'Post Time'])
            sheet.delete_rows(1, 1)
            wb.save(file_path)
    else:
        wb = Workbook()  # Creates workbook
        sheet = wb.active  # Makes workbook active
        sheet.title = 'Data'  # Renames sheet to Data
        headers = ['Title', 'Author', 'UpVotes', 'Number of Comments', 'Post URl', 'Post Time']
        sheet.append(headers)
        wb.save(excel_file_name)
    return sheet, wb, excel_file_name


add_data, save_data, file_name = create_workbook()


def chromedriver_options_headers():
    # Providing a list of user agent
    list_of_user_agents = [
        "Mozilla/5.0 (Windows NT 11.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.6998.166 Safari/537.36",
        "Mozilla/5.0 (Windows 7 Enterprise; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6099.71 Safari/537.36",
        "Mozilla/5.0 (Windows Server 2012 R2 Standard; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5975.80 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64; CentOS Ubuntu 19.04) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5957.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; WOW64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5756.197 Safari/537.36"]

    # Generating random request
    random_user_agent = random.choice(list_of_user_agents)

    # Loading Selenium options, helpfull for being undetectable

    options = Options()
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-gpu')
    options.add_argument(f"user-agent={random_user_agent}")
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument("--proxy-server='direct://'")
    options.add_argument("--proxy-bypass-list=*")
    options.add_argument('--ignore-certificate-errors')
    options.add_argument("--password-store=basic")
    options.add_argument("--no-sandbox")
    options.add_argument("--enable-automation")
    options.add_argument("--disable-browser-side-navigation")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-setuid-sandbox")
    options.add_argument("--disable-software-rasterizer")
    options.add_argument('--incognito')
    options.add_argument("--start-maximized")

    # Loading Chromedriver with options,
    web_driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options, )
    return web_driver


def scraping_post_url():
    append_url = []

    # visit Webpage
    try:
        #driver.get('https://www.reddit.com/user/leo___gangdu93/submitted/')
        driver.get('https://www.reddit.com/r/Python/')

        # This try except is written for handling pages less than given length,
        # https://www.reddit.com/user/leo___gangdu93/submitted/ has less page and no scrolling components,
        # So the code fist check for scrolling components which is available in majority of the pages
        # if not then the code finds posts available in the community.

        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[style = 'will-change: unset;']")))
            driver.execute_script("window.scrollTo(0,document.body.scrollHeight);")
            time.sleep(1)
            driver.execute_script("window.scrollTo(document.body.scrollHeight,0);")

        except TimeoutException as scroll_element_not_found:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, "//*[@class = 'w-full m-0']//*[@slot = 'full-post-link']")))
            time.sleep(2)
            logging.info('Less elements available scrolling not found', scroll_element_not_found)

        try:
            # Shows the ending scroll position of the webpage used for scrolling
            last_height = driver.execute_script("return document.body.scrollHeight")

            # Scraping Post URLS,
            # The logic behind this loop is, the site is dynamic and the index number of element changes as soon as scroll page changes.
            # So for tackling that condtion, What the code does is first picks all the href post available(Skipping promoted and Ads) and then stores it in list.
            # The duplicate href will be removed by checking the list
            # I;ve purposely not added set to remove duplicates, as set changes the postions of the data so it's difficult to test whether the scraped url is in the position of the webpage or not due to dynamic name changing
            #
            while True:

                find_length = driver.find_elements(By.XPATH, "//*[@class = 'w-full m-0']//*[@slot = 'full-post-link']")
                time.sleep(3)

                for getting_url in find_length:
                    url = getting_url.get_attribute('href')
                    if url in append_url:
                        pass
                    else:
                        append_url.append(url)
                        # print(url)

                # Here when the scrolling is done, The lenght of urls is not fixed, sometime it scrapes 28 urls and on another it scrapes 36 urls,
                # So for scraping 100 unique urls, the script scrolls almost 4 times,
                # to get 100 urls only we use list slicing below.

                if len(append_url) >= count_of_post_to_scrape:
                    del append_url[count_of_post_to_scrape:]
                    break

                # This if condition is given to handle if the posts available are less than provide value.

                if len(append_url) <= count_of_post_to_scrape:
                    try:
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        WebDriverWait(driver, 15).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "[style = 'will-change: unset;']")))
                        time.sleep(3)
                    except TimeoutException as time_out:
                        logging.error("No element found", time_out)

                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    logging.error("Less posts available and scrolling limit reached.")
                    break
                last_height = new_height

                # The above lines are for infinite scrolling till either condition matches, either count of url scraped is completed, or less number of posts are available so it will
                # scroll to end and collect all the href available

        except WebDriverException as picking_urls:
            logging.error("Except method of scraping URL", picking_urls)

    except WebDriverException as loading_website:
        logging.error("Scraping_post_function error", loading_website)

    # All the exceptions are logged in Reddit_scraper_.log
    print(len(append_url))
    return append_url


def sending_request():
    #List of user agents used to make a request randomly to the url

    list_of_user_agents = [
        "Mozilla/5.0 (Windows NT 11.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.6998.166 Safari/537.36",
        "Mozilla/5.0 (Windows 7 Enterprise; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6099.71 Safari/537.36",
        "Mozilla/5.0 (Windows Server 2012 R2 Standard; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5975.80 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64; CentOS Ubuntu 19.04) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.5957.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; WOW64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5756.197 Safari/537.36"]

    User_Agent_headers = {
        "User-Agent": random.choice(list_of_user_agents),
    }

    # Adding Random sleep while sending request to the server, IT reduces chances of being block
    try:
        sleep(randint(3, 5))
        page_content = requests.get(page_url, headers=User_Agent_headers)
        return page_content
    except Exception as error_in_request:
        logging.error('Sending request to a url ', error_in_request)


def scraping_data():
    print('Currently crawling ', page_url, count)
    # Parsing data via lxml
    soup = BeautifulSoup(site_content.text, 'lxml')


    # Scraping below data with BS4

    title = soup.find('shreddit-post')
    if 'post-title' in title.attrs:
        title_value = title['post-title']
    else:
        title_value = None

    upvote = soup.find('shreddit-post')
    if 'score' in upvote.attrs:
        upvote_text = upvote['score']
    else:
        upvote_text = None

    comment_ = soup.find('shreddit-post')
    if 'comment-count' in comment_.attrs:
        comment_count = comment_['comment-count']
    else:
        comment_count = None

    author_name = soup.find('shreddit-post-overflow-menu')
    if 'author-name' in author_name.attrs:
        author_name_val = author_name['author-name']
    else:
        author_name_val = None

    time.sleep(1)

    get_time = soup.find('faceplate-timeago')
    if 'ts' in get_time.attrs:
        get_time_val = get_time['ts']
    else:
        get_time_val = None

    clean_time = re.split('[T .]', get_time_val)

    # Creating a tuple, append it to the sheet and save the data.

    gather_data = (title_value, author_name_val, upvote_text, comment_count, page_url, clean_time[1])
    add_data.append(gather_data)
    save_data.save(file_name)


def close_driver():
    if driver:
        driver.quit()
        print('Driver closed')


if __name__ == '__main__':
    start_time = time.time()
    logging.basicConfig(filename='Reddit_scraper_.log', level=logging.ERROR,
                        format='%(asctime)s - %(levelname)s - %(message)s')

    add_data, save_data, file_name = create_workbook()
    try:
        driver = chromedriver_options_headers()
        count_of_post_to_scrape = int(input("Please enter how many post's data you want to scrape "))
        list_of_urls = scraping_post_url()
        close_driver()
        if len(list_of_urls) == 0:
            logging.error("No Url available")
        else:
            for count, page_url in enumerate(list_of_urls, 1):
                site_content = sending_request()
                if '200' in str(site_content):
                    scraping_data()
                else:
                    logging.error('Not Scraped', page_url, count)

    except WebDriverException as error:
        logging.error('_Main_ exception raised', error)
    finally:
        # close_driver()
        end_time = time.time()
        seconds = end_time - start_time
        convert_minutes = seconds / 60

        print(f"Program took {convert_minutes:.4f} minutes to run.")
